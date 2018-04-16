# !/usr/bin/python3

"""
FileName    : excel_demo.py
Author      : ken
Date        : 2018-04-16
Describe    : write and read excel
"""
import os

# 读写2007 Excel
import openpyxl
# 读写2003 excel
import xlrd
import xlwt

from read_excel.Log import MyLog as Log

currentPath = os.path.split(os.path.abspath(__file__))[0]


class ExcelSet(object):
    value = [
        ["名称", "价格", "出版社", "语言"],
        ["如何高效读懂一本书", "22.3", "机械工业出版社", "中文"],
        ["暗时间", "32.4", "人民邮电出版社", "中文"],
        ["拆掉思维里的墙", "26.7", "机械工业出版社", "中文"]
    ]

    def __init__(self, filename):
        self.logger = Log.get_log().get_logger()
        self.filename = filename
        self.files_path = None
        self.path = None

    def __set_path(self):
        self.files_path = os.path.join(currentPath, "files")
        fp = self.files_path
        if not os.path.exists(fp):
            os.mkdir(fp)

        self.path = os.path.join(fp, self.filename)

    def write_03_excel(self):
        self.__set_path()
        wb = xlwt.Workbook()
        sheet = wb.add_sheet("2003测试表")

        for i in range(len(self.value)):
            for j in range(len(self.value[i])):
                sheet.write(i, j, self.value[i][j])
        wb.save(self.path)
        print("写入数据成功！")
        self.logger.info("写入数据成功！")

    def read_03_excel(self):
        self.__set_path()
        workbook = xlrd.open_workbook(self.path)
        sheets = workbook.sheet_names()
        worksheet = workbook.sheet_by_name(sheets[0])
        for i in range(worksheet.nrows):
            for j in range(worksheet.ncols):
                print(worksheet.cell_value(i, j), "\t", end="")
            print()

    def write_07_excel(self):
        self.__set_path()
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = '2007测试表'

        for i in range(len(self.value)):
            for j in range(len(self.value[i])):
                sheet.cell(row=i + 1,
                           column=j + 1,
                           value=str(self.value[i][j]))

        wb.save(self.path)
        self.logger.info("写入数据成功！")

    def read_07_excel(self):
        self.__set_path()
        wb = openpyxl.load_workbook(self.path)
        sheets = wb.sheetnames
        worksheet = wb[sheets[0]]

        for row in worksheet.rows:
            for cell in row:
                print(cell.value, "\t", end="")
            print()


if __name__ == "__main__":
    e2003 = ExcelSet(filename="2003.xls")
    e2007 = ExcelSet(filename="2007.xlsx")

    e2003.write_03_excel()
    e2003.read_03_excel()
    print("\n")
    e2007.write_07_excel()
    e2007.read_07_excel()
